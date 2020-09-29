import wikipediaapi
import wikipedia

place_name_col_num = 1
clean_name_col_num = 2
lat_col_num = 3
long_col_num = 4


count = 0
n = 0

import openpyxl
my_folder_path = "/Users/rossmcintyre/Documents/Python/Data/"
file_name = "Destinations.xlsx"
my_path = my_folder_path + file_name
my_wb_obj = openpyxl.load_workbook(my_path)

my_sheet_obj = my_wb_obj['Sheet2']


for i in range(2, my_sheet_obj.max_row + 1):
    my_cell_obj = my_sheet_obj.cell(row = i, column = place_name_col_num)
    
    if i % 100 == 0:
        my_wb_obj.save(my_path)
        print("Saved, %d down, %d to go!" % (i-1, my_sheet_obj.max_row + 1 - i))
    

    if my_cell_obj.value:
        search = wikipedia.search(my_cell_obj.value)
       
        count = 0
        n = 0
        
        wiki_wiki = wikipediaapi.Wikipedia('en')

        while count == 0:
            n += 1

            page_py = wiki_wiki.page(search[0])

            import requests

            S = requests.Session()

            URL = "https://en.wikipedia.org/w/api.php"

            PARAMS = {
                "action": "query",
                "format": "json",
                "titles": search[n-1],
                "prop": "coordinates"
            }

            R = S.get(url=URL, params=PARAMS)

            DATA = R.json()
            PAGES = DATA['query']['pages']

            for k, v in PAGES.items():
                try:
                    print(search[0])
                    clean_name = my_sheet_obj.cell(row = i, column = clean_name_col_num)
                    clean_name.value = search[0]
                    
                    print("Latitute: " + str(v['coordinates'][0]['lat']))
                    latitude = my_sheet_obj.cell(row = i, column = lat_col_num)
                    latitude.value = str(v['coordinates'][0]['lat'])
                    
                    print("Longitude: " + str(v['coordinates'][0]['lon']))
                    longitude = my_sheet_obj.cell(row = i, column = long_col_num)
                    longitude.value = str(v['coordinates'][0]['lon'])

                    count += 1

                    
                except:
                    if n == 2:
                        clean_name = my_sheet_obj.cell(row = i, column = clean_name_col_num)
                        clean_name.value = "Error"
                        
                        latitude = my_sheet_obj.cell(row = i, column = lat_col_num)
                        latitude.value = "Error"

                        longitude = my_sheet_obj.cell(row = i, column = long_col_num)
                        longitude.value = "Error"

                        
                        count += 1
                    continue
        
        else:
            continue

my_wb_obj.save(my_path)
